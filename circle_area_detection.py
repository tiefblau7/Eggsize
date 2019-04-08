import cv2
import numpy as np
import copy
import imutils
import re
from datetime import datetime
import matplotlib.pyplot as plt
import openpyxl
from PIL import Image

today = datetime.now()
str_today = today.strftime('%Y%m%d%H%M%S')
#読み込む画像の名前
image_file = "input/P1180025.JPG"
#P1180025は226個（手計測）
img = cv2.imread(image_file)
img_gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
img_gray = cv2.medianBlur(img_gray,5)

if len(img.shape) == 3:
    height, width, channels = img.shape[:3]
else:
    height, width = img.shape[:2]
    channels = 1

print("width: " + str(width))
print("height: " + str(height))
print("channels: " + str(channels))
print("dtype: " + str(img.dtype))

if "/" in image_file:
    m = re.search('/([^/]*)(\..*)', image_file)
else:
    m = re.search('(.*)(\..*)', image_file)
#originalのまま保存するとき
#filename = "output/"+str(m.group(1) + "_" + str_today + m.group(2))
#tiff形式で保存するとき
filename = "output/"+str(m.group(1) + "_" + str_today + ".tif")
cv2.namedWindow('img', cv2.WINDOW_NORMAL)

def nothing(x):
    pass

def circledraw(img1,x1,y1,rad1):
    #　円を描く
    #cv2.circle(画像, 中心座標, 半径, 色, 線の太さ)
    cv2.circle(img1,(x1,y1),rad1,(0,255,0),6)
    # 円の中心を描く。
    cv2.circle(img1,(x1,y1),5,(0,0,255),3)

cv2.createTrackbar("xaxis", 'img', width//2, width, nothing)
cv2.createTrackbar("yaxis", 'img', height//2, height, nothing)
cv2.createTrackbar("rad1", 'img', height//3, height, nothing)

while(1):
    img_orig = copy.deepcopy(img)
    x = cv2.getTrackbarPos("xaxis", 'img')
    y = cv2.getTrackbarPos("yaxis", 'img')
    rad = cv2.getTrackbarPos("rad1", 'img')
    circledraw(img_orig, x, y, rad)
    cv2.imshow('img', img_orig)
    if cv2.waitKey(1000) & 0xFF == ord("q"):
        break
print(f"x座標,y座標,半径:{x},{y},{rad}")
cv2.destroyAllWindows()

rad2 = rad**2

#---以下、上記で検出した円の内部における卵の検出および表示---
cv2.namedWindow('detect_img', cv2.WINDOW_NORMAL)
cv2.namedWindow('thresh_img', cv2.WINDOW_NORMAL)
cv2.createTrackbar("Threshold", 'detect_img',100,255,nothing)
while(1):
    img_orig = copy.deepcopy(img)
    thr = cv2.getTrackbarPos("Threshold", 'detect_img')
    ret, thresh = cv2.threshold(img_gray, thr, 255, cv2.THRESH_BINARY)
    image, contours, hierarchy = cv2.findContours(thresh, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    cnt = 0
    contours1 = copy.deepcopy(contours)

    for i in range(len(contours)):
        contours[i] = contours[i].reshape([-1,2])
        lencon = len(contours[i])
        for j in range(lencon):
            xz = contours[i][j][0]
            yz = contours[i][j][1]
            if (xz - x)**2 + (yz - y)**2 > rad2:
                #np.delete(contours1, i - cnt, 0)
                contours1.pop(i - cnt)
                cnt += 1
                break

    img_cnt = cv2.drawContours(img_orig, contours1, -1, (0,255,0), 2)
    circledraw(img_orig, x, y, rad)
    cv2.imshow('detect_img', img_cnt)
    cv2.imshow('thresh_img', thresh)
    if cv2.waitKey(1000) & 0xFF == ord("q"):
        break

#thr170で247contours検出（範囲指定無し）
#contours[i].reshape([-1,2])[0][0]
cv2.destroyAllWindows()
cont_area_list = []
for conts in contours1:
    cont_area_list.append(cv2.contourArea(conts))
arr1 = np.array(cont_area_list)

dif = (np.percentile(arr1, 70) - np.median(arr1))*2
med = np.median(arr1)
#arr2 = arr1[(med-dif<arr1)&(arr1 < med+dif)]
#arr2 = arr1[(1000<arr1)&(arr1 < 4000)]
arr2 = arr1[(80<arr1)&(arr1 < 300)]
print(f"中央値は{med}、75 percentileまでの幅は{dif}")
print(arr2)
#print(cont_area_list)
print(f"削除した領域は：{cnt}個")
print(f"残った領域は：{len(arr2)}")
img_cnt = imutils.resize(img_cnt, width=480)
thresh = imutils.resize(thresh, width=480)
cv2.imwrite("excel_data/detect_img.png", img_cnt)
cv2.imwrite("excel_data/detect_img2.png", thresh)

#print(f"最大値は{np.amax(arr1)}です。")
#print(arr1)

#ヒストグラム描写
arr1 = np.where(arr1<500, arr1, 500)
plt.xlabel('pixel(over 500 as 500)')
plt.ylabel('number of eggs')
plt.hist(arr1, range=(0, 500), bins=50)
plt.savefig("excel_data/histplot.png", dpi = 150)
plt.show()


#Excelシート書き込み
hist_img = cv2.imread("excel_data/histplot.png")
hist_img = imutils.resize(hist_img, width = 480)
cv2.imwrite("excel_data/histplot.png", hist_img)

book = openpyxl.load_workbook("excel_data/eggsize_org.xlsx")
sheet = book['Sheet1']
hist_img = openpyxl.drawing.image.Image("excel_data/histplot.png")
book.worksheets[0].add_image(hist_img, "F2")
cd_img = openpyxl.drawing.image.Image("excel_data/detect_img.png")
book.worksheets[0].add_image(cd_img, "F16")
cd_img2 = openpyxl.drawing.image.Image("excel_data/detect_img2.png")
book.worksheets[0].add_image(cd_img2, "K16")

cnt = 1
for pix in arr2:
    cnt += 1
    sheet.cell(row=cnt, column=2).value = pix
sheet.cell(row=2, column=4).value = len(arr2)

savename = str("excel_data/" + m.group(1) + "_" + str_today + ".xlsx")
book.save(savename)
#print(np.delete(contours, -1, 0))
#print(len(np.delete(contours[1], 0, 0)))
#(contours[1], 0, 0)について、1つ目の引数がどのarrayか、2つ目の引数がそのarrayの何番目のものか、3つ目の因数がaxisのことで、0始まり。2次元行列の場合0なら行だし1だと列になる。
#cv2.imwrite(filename, th3)


#Thresholdは透過光のものを見て一律◯◯と決める。要マニュアルフォーカス。
#卵として認める大きさは実際に上の条件ででたヒストグラム等を元にこれも一律に固定する。
#チタンの細線ではなく、マイクロメーター（顕微鏡の校正機）を用いてピクセル：長さを出す。
